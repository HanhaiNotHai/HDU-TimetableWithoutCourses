"""无课表自动生成"""


def main():
    """main"""
    # os listdir遍历文件夹 chdir切换工作目录
    import os
    import re
    # sys.argv[0]得到脚本文件名
    import sys

    # excel工作簿 单元格 样式 工作表
    from openpyxl import Workbook
    from openpyxl.cell.cell import Cell
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.worksheet.worksheet import Worksheet
    # 读取解析pdf内容
    import pdfplumber

    # 基础变量
    # 改变工作目录到脚本所在目录
    os.chdir(os.path.split(os.path.abspath(sys.argv[0]))[0])

    # '课表'目录
    timetables_dir = '课表'

    def not_digit_index(s: str) -> int:
        """s中第一个非数字的下标"""
        for i, c in enumerate(s):
            if not c.isdigit():
                return i

    # 部门目录
    departments_dirs = os.listdir(timetables_dir)
    # 部门目录名第一个非数字的下标
    departments_index = {department: not_digit_index(department)
                         for department in departments_dirs}
    # 部门目录按开头数字排序
    departments_dirs.sort(key=lambda department: int(order)
                          if (order := department[:departments_index[department]]) else 999)
    # 部门名
    departments = [department[departments_index[department]:]
                   for department in departments_dirs]
    # 部门数量
    departments_num = len(departments)
    # 将部门目录转换成路径
    departments_dirs = [os.path.join(timetables_dir, departments_dir)
                        for departments_dir in departments_dirs]

    def get_weeks_num() -> int:
        """获取总周数
        不同周数出现的次数，出现6次返回
        """
        weeks_num = {week: -5 for week in range(16, 19)}
        for filepath, _, filenames in os.walk(timetables_dir):
            for filename in filenames:
                with pdfplumber.open(os.path.join(filepath, filename)) as pdf:
                    for page in pdf.pages:
                        table = page.extract_table()
                        for row in table[2:]:
                            for info in row[2:]:
                                if info:
                                    for week in weeks_num:
                                        if occur_times := info.count(str(week) + '周'):
                                            weeks_num[week] += occur_times
                                            if weeks_num[week]:
                                                return week
        return int(input('请输入本学期总周数:'))

    # 总周数
    weeks_num = get_weeks_num()

    # 提取上课节次和周数
    re_class_time = re.compile(''.join([
        r'\(([1-9]|1[0-2])-([1-9]|1[0-2])节\)(([1-9]|1[0-',
        str(weeks_num % 10),
        r'])(-([1-9]|1[0-',
        str(weeks_num % 10), r']))?周(\((单|双)\))?,?)+'
    ]))

    # excel表行列变量
    # 输出无课表中每节课开始的行
    start_row = list(range(3, 3 + 8 * departments_num, departments_num))
    # 下午向下移一行
    start_row[3:6] = [row+1 for row in start_row[3:6]]
    # 晚上向下移两行
    start_row[6:] = [row+2 for row in start_row[6:]]
    # 输出无课表中每节课结束的行
    end_row = [row + departments_num - 1 for row in start_row]

    # 课表中节次开始行数
    start_row_class_timetable = [2, 4, 6, 7, 9, 10, 11, 13]
    # 节次对应个人无课表行数
    class2table = [-1, 0, 0, 1, 1, 2, 3, 3, 4, 5, 6, 6, 7]

    # 新建工作簿wb 工作表ws
    wb = Workbook()
    ws: Worksheet = wb.active

    # 设置ws excel框架 样式
    # 所有单元格 设为字符串 居中 自动换行 边框 字体
    side = Side(border_style='medium')
    for i in range(1, end_row[7] + 1):
        for j in range(1, 9):
            cell: Cell = ws.cell(i, j, '')
            cell.alignment = Alignment(horizontal='center',
                                       vertical='center',
                                       wrap_text=True)
            cell.border = Border(side, side, side, side)
            cell.font = Font(size=12)

    # 合并单元格
    ws.merge_cells(None, 1, 1, 1, 8)
    # 设置内容、格式
    ws.cell(1, 1, '无课表').font = Font(size=14)
    # 设置高度、宽度
    ws.row_dimensions[1].height = 43

    ws.merge_cells(None, 2, 1, 2, 2)
    ws.cell(2, 1, '时间')
    ws.cell(2, 3, '部门')
    week_name = ['星期一', '星期二', '星期三', '星期四', '星期五']
    for col, week in enumerate(week_name, 4):
        ws.cell(2, col, week)
    ws.row_dimensions[2].height = 36
    for i in range(3):
        ws.column_dimensions[chr(ord('A') + i)].width = 11
    for i in range(3, 8):
        ws.column_dimensions[chr(ord('A') + i)].width = 33

    ws.merge_cells(None, start_row[0], 1, end_row[2], 1)
    ws.cell(start_row[0], 1, '上午')
    ws.merge_cells(None, start_row[3], 1, end_row[5], 1)
    ws.cell(start_row[3], 1, '下午')
    ws.merge_cells(None, start_row[6], 1, end_row[7], 1)
    ws.cell(start_row[6], 1, '晚上')

    # 上午、下午、晚上中间的空行
    ws.merge_cells(None, end_row[2] + 1, 1, end_row[2] + 1, 8)
    ws.merge_cells(None, end_row[5] + 1, 1, end_row[5] + 1, 8)

    time = ['1-2', '3-4', '5', '6-7', '8', '9', '10-11', '12']
    for i in range(8):
        ws.merge_cells(None, start_row[i], 2, end_row[i], 2)
        ws.cell(start_row[i], 2, time[i])

        # 部门列
        for department_index, department in enumerate(departments):
            ws.cell(start_row[i] + department_index, 3, department)

    def start_end_week2str(start_week: int, end_week: int) -> str:
        ret = [str(start_week)]
        if end_week > start_week:
            if end_week > start_week + 1:
                # 连续2周以上用-连接
                ret.append('-')
            else:
                # 连续2周用、连接
                ret.append('、')
            ret.append(str(end_week))
        return ''.join(ret)

    # 读取pdf课表信息 建立无课表
    for department_index, department in enumerate(departments):
        print(department)
        # 部门目录
        department_dir = departments_dirs[department_index]
        # 部门内所有文件按开头数字排序
        files = sorted(os.listdir(department_dir),
                       key=lambda x: int(order) if (order := x[:not_digit_index(x)]) else 999)
        for file in files:
            # 读取pdf
            with pdfplumber.open(os.path.join(department_dir, file)) as pdf:
                # pdf每页提取表格
                tables: list[list[list[str | None]]] = [table for page in pdf.pages
                                                        if (table := page.extract_table())]
                # 读取姓名
                name = tables[0][0][0].split('课表')[0]
                print('\t', name)

                # 将多页的表格合并成一个表格
                for page in range(len(tables) - 1):
                    # 后一页还是课表内容 不是“其他课程...”、“：讲课学时...”行
                    if len(tables[page + 1][0]) == 9:
                        # 前一页最后一行或者后一页第一行节次单元格为'',要合并
                        if tables[page][-1][1] == '' or tables[page + 1][0][1] == '':
                            for col in range(9):
                                # 后一页第一行单元格有内容
                                if tables[page + 1][0][col]:
                                    # 找到内容应合并到的前一页的单元格
                                    row_off = -1
                                    while tables[page][row_off][col] is None:
                                        row_off -= 1
                                    # 将后一页内容合并到前一页
                                    tables[page][row_off][col] = ''.join([
                                        tables[page][row_off][col], tables[page + 1][0][col]
                                    ])
                            # 已合并,删除后一页第一行
                            tables[page + 1] = tables[page + 1][1:]
                # 个人课表
                timetable = [row for table in tables for row in table]

                # 建立个人无课表
                timetable_without_courses = [[set(range(1, weeks_num + 1))
                                              for _ in range(5)] for _ in range(8)]
                for day in range(5):
                    for _class in range(8):
                        # 个人课表单元格有内容
                        if cur_cell := timetable[start_row_class_timetable[_class]][day + 2]:
                            # 删除\n
                            cur_cell = cur_cell.replace('\n', '')
                            # 上课节次 周数
                            class_times = re_class_time.finditer(cur_cell)
                            for class_time in class_times:
                                # (8-9节)1-7周(单),8-11周,13-16周
                                class_time = class_time.group().split('节)')
                                # 上课开始 结束节次
                                start_class, end_class = (
                                    int(_class) for _class in class_time[0][1:].split('-')
                                )
                                # 上课起止周数
                                start_end_weeks = class_time[1].split(',')
                                for start_end_week in start_end_weeks:
                                    start_end_week = start_end_week.split('周')
                                    # 是否单双周
                                    mark = start_end_week[1]
                                    start_end_week = \
                                        start_end_week[0].split('-')

                                    if len(start_end_week) == 1:
                                        # 单独一周
                                        start_week = int(start_end_week[0])
                                        end_week = start_week
                                    else:
                                        # 有起止周
                                        start_week, end_week = (int(week)
                                                                for week in start_end_week)

                                    if mark != '':
                                        # (单) (双) 去除括号
                                        mark = mark[1]
                                        # 起止周次不符合单双
                                        if mark == '单':
                                            if start_week % 2 == 0:
                                                start_week += 1
                                            if end_week % 2 == 0:
                                                end_week -= 1
                                        elif mark == '双':
                                            if start_week % 2:
                                                start_week += 1
                                            if end_week % 2:
                                                end_week -= 1

                                    # 有课的周数
                                    weeks_with_courses = set(range(start_week, end_week + 1,
                                                                   2 if mark else 1))
                                    for __class in range(class2table[start_class],
                                                         class2table[end_class] + 1):
                                        timetable_without_courses[__class][day] -=\
                                            weeks_with_courses

                # 将个人无课表写入ws中
                for day in range(5):
                    for _class in range(8):
                        # 无序set转为有序list
                        cur_cell = sorted(list(
                            timetable_without_courses[_class][day]
                        ))
                        # 个人无课表单元格有内容
                        if cur_cell:
                            cell: Cell = ws.cell(start_row[_class] + department_index,
                                                 day + 4)
                            cell.value = ''.join([cell.value, '，', name])
                            # 不是所有周都无课
                            if len(cur_cell) < weeks_num:
                                value = []
                                start_week = end_week = cur_cell[0]
                                for week_index in range(1, len(cur_cell)):
                                    if cur_cell[week_index] == cur_cell[week_index - 1] + 1:
                                        # 连续无课
                                        end_week = cur_cell[week_index]
                                    else:
                                        value.append(start_end_week2str(start_week,
                                                                        end_week))
                                        value.append('、')
                                        start_week = end_week = cur_cell[week_index]
                                value.append(start_end_week2str(start_week,
                                                                end_week))
                                cell.value = ''.join([cell.value,
                                                      ''.join(value)])

    # 去除单元格开头'，'
    for row in range(3, end_row[7]):
        for col in range(4, 9):
            cell = ws.cell(row, col)
            if cell.value and cell.value[0] == '，':
                cell.value = cell.value[1:]

    # 保存excel到 无课表.xlsx
    wb.save('无课表.xlsx')


if __name__ == '__main__':
    try:
        # 打印异常信息
        from traceback import print_exc
        main()
    except:
        print_exc()
        print(''.join(['！！！出错了，请反馈！！！\n' * 3,
              'qq 1932232849 或者github issues']))
    finally:
        input('按回车键退出')
